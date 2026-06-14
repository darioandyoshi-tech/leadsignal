"""Nebraska Department of Revenue delinquent real property tax list.

The DOR publishes county-level Excel files each year at:
  https://revenue.nebraska.gov/PAD/real-property/nebraska-delinquent-real-property-list

Douglas County 2026 example:
  https://revenue.nebraska.gov/sites/default/files/doc/pad/delinquent_real_prop/2026/28Douglas_delinq2026.xlsx
"""

import io
import re
import requests
from datetime import datetime
from typing import List

import openpyxl

from scraper.config import DOR_DELINQ_MIN_BALANCE, DOR_DELINQ_COUNTIES
from scraper.db_client import get_or_create_company, insert_signal, Session, signal_exists
from app.models import SignalType


BASE_PAGE = "https://revenue.nebraska.gov/PAD/real-property/nebraska-delinquent-real-property-list"
FILE_PATTERN = re.compile(r'href="([^"]+\.xlsx)"', re.I)


def _discover_excel_urls(year: int = None) -> List[str]:
    """Find xlsx links on the DOR delinquency page for the requested year."""
    year = year or datetime.utcnow().year
    try:
        r = requests.get(BASE_PAGE, timeout=30)
        r.raise_for_status()
    except Exception:
        return []
    links = FILE_PATTERN.findall(r.text)
    absolute = []
    for link in links:
        if link.startswith("http"):
            absolute.append(link)
        elif link.startswith("/"):
            absolute.append(f"https://revenue.nebraska.gov{link}")
        else:
            absolute.append(f"https://revenue.nebraska.gov/PAD/real-property/{link}")
    return [u for u in absolute if str(year) in u]


def _download_excel(url: str):
    try:
        r = requests.get(url, timeout=60)
        r.raise_for_status()
        return openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    except Exception:
        return None


def _money_value(cell) -> int:
    if cell is None:
        return 0
    v = cell.value
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace("$", "").replace(",", "").strip()
    try:
        return int(float(s))
    except ValueError:
        return 0


def _str_value(cell) -> str:
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def _parse_county(workbook: openpyxl.Workbook, min_balance: int, source_url: str) -> dict:
    created = 0
    skipped = 0
    inspected = 0

    sheet = workbook.active
    headers = {}
    header_row = None
    for i, row in enumerate(sheet.iter_rows(max_row=20), start=1):
        vals = [_str_value(c) for c in row]
        lower = [v.lower() for v in vals]
        if any("owner" in v for v in lower) and any(
            "delinquent" in v or "amount" in v or "balance" in v for v in lower
        ):
            headers = {v.lower(): idx for idx, v in enumerate(vals) if v}
            header_row = i
            break

    if not headers:
        return {
            "source": "ne_dor_delinquency",
            "signals_created": 0,
            "signals_skipped": 0,
            "inspected": 0,
            "note": "Could not locate header row",
        }

    owner_idx = headers.get("owner") or next(
        (i for i, h in headers.items() if "owner" in i), None
    )
    parcel_idx = headers.get("parcel") or next(
        (i for i, h in headers.items() if "parcel" in i or "pin" in i), None
    )
    address_idx = headers.get("property address") or next(
        (i for i, h in headers.items() if "address" in i or "location" in i), None
    )
    balance_idx = headers.get("total delinquent amount") or next(
        (i for i, h in headers.items() if "delinquent" in i or "amount" in i or "balance" in i or "due" in i),
        None,
    )
    city_idx = headers.get("city")
    zip_idx = headers.get("zip")

    for row in sheet.iter_rows(min_row=header_row + 1):
        inspected += 1
        owner = _str_value(row[owner_idx]) if owner_idx is not None else "Unknown Owner"
        parcel = _str_value(row[parcel_idx]) if parcel_idx is not None else ""
        address = _str_value(row[address_idx]) if address_idx is not None else ""
        city = _str_value(row[city_idx]) if city_idx is not None else "Omaha"
        zip_code = _str_value(row[zip_idx]) if zip_idx is not None else ""
        balance = _money_value(row[balance_idx]) if balance_idx is not None else 0

        if balance < min_balance:
            continue
        if "omaha" not in city.lower() and not any(
            z.startswith("681") for z in [zip_code, address]
        ):
            continue

        with Session() as session:
            company = get_or_create_company(
                session,
                owner or "Unknown Owner",
                city=city,
                state="Nebraska",
                zip_code=zip_code or None,
                external_ids={"dor_parcel": parcel} if parcel else {},
            )
            headline = f"Tax-delinquent property: {address or parcel} (${balance:,.0f})"
            if signal_exists(session, company.id, SignalType.tax_delinquency, headline):
                skipped += 1
                continue
            sid = insert_signal(
                company_id=company.id,
                signal_type=SignalType.tax_delinquency,
                severity=3 if balance >= 5000 else 2,
                headline=headline,
                summary=f"Owner: {owner}\nParcel/PIN: {parcel}\nAddress: {address}, {city} {zip_code}\nDelinquent Balance: ${balance:,.0f}",
                source_url=source_url,
                source_api="ne_dor_delinquency",
                location_name=f"{address}, {city} {zip_code}" if address else city,
                metadata={
                    "parcel": parcel,
                    "address": address,
                    "city": city,
                    "zip": zip_code,
                    "delinquent_balance": balance,
                },
                session=session,
            )
            if sid:
                created += 1
            session.commit()

    return {
        "source": "ne_dor_delinquency",
        "signals_created": created,
        "signals_skipped": skipped,
        "inspected": inspected,
    }


def run(year: int = None) -> dict:
    year = year or datetime.utcnow().year
    urls = _discover_excel_urls(year)
    if not urls:
        urls = [
            f"https://revenue.nebraska.gov/sites/default/files/doc/pad/delinquent_real_prop/{year}/28Douglas_delinq{year}.xlsx"
        ]

    total = {
        "source": "ne_dor_delinquency",
        "signals_created": 0,
        "signals_skipped": 0,
        "inspected": 0,
        "urls": urls,
    }
    for url in urls:
        wb = _download_excel(url)
        if wb is None:
            continue
        result = _parse_county(wb, DOR_DELINQ_MIN_BALANCE, url)
        total["signals_created"] += result["signals_created"]
        total["signals_skipped"] += result["signals_skipped"]
        total["inspected"] += result["inspected"]
        total["note"] = result.get("note")
        wb.close()

    return total


if __name__ == "__main__":
    print(run())
