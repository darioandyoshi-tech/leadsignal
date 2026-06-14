"""Nebraska Liquor Control Commission active license roster.

URL: https://lcc.nebraska.gov/licensing-sdl/active-license-roster
Excel link pattern: /sites/default/files/licensing/Active Licensing/Active Roster {date}.xlsx
"""

import io
import re
import requests
from datetime import datetime
from typing import List

import openpyxl

from scraper.config import OMAHA
from scraper.db_client import get_or_create_company, insert_signal, Session, signal_exists
from app.models import SignalType


BASE_URL = "https://lcc.nebraska.gov/licensing-sdl/active-license-roster"
LINK_RE = re.compile(r'href="([^"]+Active%20Roster[^"]+\.xlsx)"')


def _discover_roster_url() -> str | None:
    try:
        r = requests.get(BASE_URL, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
    except Exception:
        return None
    links = LINK_RE.findall(r.text)
    if not links:
        return None
    link = links[0]
    if link.startswith("http"):
        return link
    return f"https://lcc.nebraska.gov{link}"


def _download_workbook(url: str):
    try:
        r = requests.get(url, timeout=60, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        return openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    except Exception:
        return None


def _is_omaha_area(city: str, county: str, address: str) -> bool:
    haystack = f"{city} {county} {address}".lower()
    return "omaha" in haystack or any(z in address for z in OMAHA.zips) or "douglas" in haystack or "sarpy" in haystack


def run() -> dict:
    created = 0
    skipped = 0
    inspected = 0

    url = _discover_roster_url()
    if not url:
        return {"source": "nlcc_licenses", "signals_created": 0, "signals_skipped": 0, "inspected": 0, "error": "Could not discover roster URL"}

    wb = _download_workbook(url)
    if not wb:
        return {"source": "nlcc_licenses", "signals_created": 0, "signals_skipped": 0, "inspected": 0, "error": "Could not download workbook", "url": url}

    ws = wb.active
    headers = None
    header_row = None
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
        if row and "License Number" in str(row[4]):
            headers = [str(c).strip().lower() if c else "" for c in row]
            header_row = i
            break

    if not headers:
        wb.close()
        return {"source": "nlcc_licenses", "signals_created": 0, "signals_skipped": 0, "inspected": 0, "error": "Could not locate header row", "url": url}

    idx = {name: i for i, name in enumerate(headers)}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        inspected += 1
        trade_name = str(row[idx.get("trade name", 6)] or "").strip()
        licensee = str(row[idx.get("licensee", 14)] or "").strip()
        address = str(row[idx.get("address", 7)] or "").strip()
        city = str(row[idx.get("city", 8)] or "").strip()
        county = str(row[idx.get("county", 10)] or "").strip()
        license_num = str(row[idx.get("license number", 4)] or "").strip()
        license_class = str(row[idx.get("class", 2)] or "").strip()
        status = str(row[idx.get("license state", 5)] or "").strip()
        effective = row[idx.get("current effective date", 17)]
        expiration = row[idx.get("current expiration date", 18)]

        if status.lower() != "active":
            continue
        if not _is_omaha_area(city, county, address):
            continue

        name = trade_name or licensee or "Unknown Licensee"
        with Session() as session:
            company = get_or_create_company(
                session,
                name,
                city=city or "Omaha",
                state="Nebraska",
                external_ids={"nlcc_license": license_num},
            )
            headline = f"Active liquor license: {name} ({license_class})"
            if signal_exists(session, company.id, SignalType.business_license, headline):
                skipped += 1
                continue
            sid = insert_signal(
                company_id=company.id,
                signal_type=SignalType.business_license,
                severity=2,
                headline=headline,
                summary=f"Licensee: {licensee}\nTrade Name: {trade_name}\nAddress: {address}, {city}, {county} County\nLicense #: {license_num}\nClass: {license_class}\nEffective: {effective}\nExpiration: {expiration}",
                source_url=url,
                source_api="nlcc_licenses",
                location_name=f"{address}, {city}" if address else city,
                metadata={
                    "license_number": license_num,
                    "license_class": license_class,
                    "licensee": licensee,
                    "trade_name": trade_name,
                    "address": address,
                    "city": city,
                    "county": county,
                    "effective_date": str(effective) if effective else None,
                    "expiration_date": str(expiration) if expiration else None,
                },
                session=session,
            )
            if sid:
                created += 1
            session.commit()

    wb.close()
    return {"source": "nlcc_licenses", "signals_created": created, "signals_skipped": skipped, "inspected": inspected, "url": url}


if __name__ == "__main__":
    print(run())
